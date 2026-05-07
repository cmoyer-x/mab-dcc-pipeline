"""
generate_excel_outputs.py — Snakemake script
Generate final Excel spreadsheets.
"""
import csv
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

abs_csv   = snakemake.input.abs_csv
mas_csv   = snakemake.input.mas_csv
pairs_csv = snakemake.input.pairs_all
out_dcc   = snakemake.output.dcc_final
out_trans = snakemake.output.trans_final

dcc_colors = {
    'DCC1':'AED6F1','DCC2':'A9DFBF','DCC3':'FAD7A0','DCC4':'D7BDE2',
    'DCC5':'F9E79F','DCC6':'FADBD8','DCC7':'D5F5E3',
    'Non-DCC':'F5CBA7','Unknown':'FFFFFF'
}

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

all_isolates = []
for fpath in [abs_csv, mas_csv]:
    try:
        with open(fpath) as f:
            all_isolates.extend(list(csv.DictReader(f)))
    except:
        pass

pairs = []
try:
    with open(pairs_csv) as f:
        pairs = list(csv.DictReader(f))
except:
    pass

# DCC assignments workbook
wb = Workbook()
ws1 = wb.active
ws1.title = "All Isolates"

if all_isolates:
    headers = list(all_isolates[0].keys())
    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color='2C3E50')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    ws1.row_dimensions[1].height = 30
    for ri, row in enumerate(all_isolates, 2):
        color = dcc_colors.get(row.get('DCC',''), 'FFFFFF')
        for ci, h in enumerate(headers, 1):
            cell = ws1.cell(row=ri, column=ci, value=row.get(h,''))
            cell.font = Font(name='Arial', size=9)
            cell.fill = PatternFill('solid', start_color=color)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

ws2 = wb.create_sheet("Summary")
ws2['A1'] = 'DCC Assignment Summary'
ws2['A1'].font = Font(bold=True, size=13, name='Arial')
counts = Counter(r.get('DCC','Unknown') for r in all_isolates)
for ri, (dcc, n) in enumerate(sorted(counts.items()), 3):
    ws2.cell(row=ri, column=1, value=dcc).font = Font(name='Arial', size=10)
    ws2.cell(row=ri, column=2, value=n).font = Font(name='Arial', size=10)

wb.save(out_dcc)

# Transmission pairs workbook
wb2 = Workbook()
ws = wb2.active
ws.title = "Transmission Pairs"

if pairs:
    headers2 = list(pairs[0].keys())
    for ci, h in enumerate(headers2, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color='2C3E50')
        cell.border = border
    for ri, row in enumerate(pairs, 2):
        for ci, h in enumerate(headers2, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(h,''))
            cell.font = Font(name='Arial', size=9)
            cell.border = border

wb2.save(out_trans)
print(f"Excel written: {len(all_isolates)} isolates, {len(pairs)} pairs")
